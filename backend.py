import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.merge import MergeCells

import os
import time

# Пути к файлам
spec_file = "/Users/vladk/Downloads/Telegram Desktop/ЮМП.250.212.045.03 Спецификация.xlsx"
ekb_file = "/Users/vladk/Downloads/Telegram Desktop/список паспартов ЭКБ.xlsx"
output_path = "/Users/vladk/Downloads/Telegram Desktop/output/merged_output.xlsx"

def load_data(spec_file, ekb_file):
    """Загрузка данных из всех листов файла спецификации."""
    all_sheets = pd.read_excel(spec_file, sheet_name=None)  # Загружаем все листы
    specification = pd.concat(all_sheets.values(), ignore_index=True)  # Объединяем в один DataFrame
    passports = pd.read_excel(ekb_file, sheet_name='Лист1', dtype={'Дата': str})  # Дата как текст
    return specification, passports

def prepare_data(specification, passports):
    """Очистка и подготовка данных."""
    specification = specification.dropna(subset=['Наименование'])

    # Удаляем дубликаты по 'Наименование'
    specification = specification.drop_duplicates(subset=['Наименование'], keep='first')

    passports.columns = ['Наименование', 'Паспорт', 'Дата']

    # Преобразуем дату в строку (если она не строка)
    passports['Дата'] = passports['Дата'].astype(str)
    passports['Дата'] = fix_date_format(passports['Дата'])


    return specification, passports

def fix_date_format(date_series):
    """
    Приводит даты в формате 'M.YYYY' или 'M.YY' к единому формату 'MM.YYYY'
    Примеры:
    '5.2021' -> '05.2021'
    '1.202' -> '01.2020'
    '12.202' -> '12.2020'
    """
    def process_date(date_str):
        if pd.isna(date_str):
            return date_str
        
        # Если значение уже является строкой в нужном формате, пропускаем
        if isinstance(date_str, str) and len(date_str.split('.')) == 2:
            month, year = date_str.split('.')
            
            # Добавляем ведущий ноль к месяцу, если нужно
            month = month.zfill(2)
            
            # Исправляем год (добавляем 0 если год трехзначный)
            year = year.ljust(4, '0') if len(year) == 3 else year
            
            return f"{month}.{year}"
        return date_str
    
    return date_series.apply(process_date)

def extract_year_and_add_25(date_str):
    """Извлекает год из строки формата 'MM.YYYY' и прибавляет 25."""
    try:
        year = int(date_str.split('.')[-1])  # Берем только год
        return year + 25  # Прибавляем 25 и конвертируем в строку
    except (ValueError, IndexError):
        return ""  # Если ошибка, оставляем пустым

def merge_data(specification, passports):
    """Объединение данных по наименованию."""
    merged_data = pd.merge(specification, passports, on='Наименование', how='left')

    # Преобразуем столбец "Дата" в строку с явным форматом MM.YYYY
    merged_data['Дата'] = merged_data['Дата'].apply(lambda x: str(x) if pd.notna(x) else '').str[:7]

    # Если "Паспорт" пустой, то "H" остается пустым, иначе заполняем 25
    merged_data['H'] = merged_data['Паспорт'].apply(lambda x: 25 if pd.notna(x) and x != '' else '')

    # Вычисляем "I" только если есть "Дата"
    merged_data['I'] = merged_data['Дата'].apply(lambda x: extract_year_and_add_25(x) if x else "")

    return merged_data



def create_result_table(merged_data):
    """Создание результирующей таблицы с пустыми столбцами и нумерацией."""
    result = pd.DataFrame({
        'A': merged_data['Поз.']-1,
        'B': '',
        'C': merged_data['Наименование'],
        'D': merged_data.get('Кол.', ''),
        'E': merged_data.get('Кол.', ''),
        'F': merged_data['Паспорт'],
        'G': merged_data['Дата'],
        'H': merged_data['H'],
        'I': merged_data['I']
    })
    # result.insert(0, '№', range(1, len(result) + 1))
    return result

def add_section_names(result, specification):
    """Добавление названий разделов и разбиение длинных наименований."""
    section_names = specification[specification['Наименование'].str.contains('Конденсаторы|Микросхемы|Диоды|Транзисторы', case=False, na=False)]['Наименование']
    final_data = []

    for _, row in result.iterrows():
        name = row['C']
        if name in section_names.values:
            final_data.append(['', '', name, '', '', '', '', '', ''])
        else:
            words = name.split()  # Разбиваем на слова
            lines = []
            current_line = ""

            for word in words:
                if len(current_line) + len(word) + 1 <= 18:  # +1 для пробела
                    current_line += (" " if current_line else "") + word
                else:
                    lines.append(current_line)
                    current_line = word

            if current_line:
                lines.append(current_line)

            # Заполняем строки
            for i, line in enumerate(lines):
                if i == 0:
                    final_data.append([row['A'], "", line, row['D'], row['E'], row['F'], row['G'], row['H'], row['I']])
                else:
                    final_data.append(["", "", line, "", "", "", "", "", ""])  # Пустая строка с текстом
    return final_data



from openpyxl.styles import Font, Alignment

def save_to_excel(final_data, output_path):
    """Сохранение данных в Excel с обработкой ошибки доступа."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    row_count = 0
    sheet_number = 1

    italic_sections = [
        "Конденсаторы", "Микросхемы", "Катушки", "индуктивности", "Резисторы",
        "Печатная плата", "Транзисторы", "Диоды", "Соединения", "контактные"
    ]

    for row in final_data:
        if row_count >= 18:
            sheet_number += 1
            ws = wb.create_sheet(title=f"Лист{sheet_number}")
            row_count = 0

        ws.append(row)
        row_count += 1

    # Применение стилей ко всем листам
    for sheet in wb.worksheets:
        # Установка формата текста для столбцов G, H, I
        for col in sheet.iter_cols(min_col=7, max_col=9):
            for cell in col:
                cell.number_format = '@'

        # Установка ширины столбцов
        column_widths = [63, 276, 255, 80, 80, 265, 82, 99, 99]
        columns = ["A", "B", "C", "D", "E", "F", "G", "H", "I"]

        for col, width in zip(columns, column_widths):
            sheet.column_dimensions[col].width = width / 13.43

        # Курсив для названий разделов
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
            for cell in row:
                if cell.value and any(section in str(cell.value) for section in italic_sections):
                    cell.font = Font(italic=True)

        # Установка размера шрифта 8 для столбца G
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=7):  # Столбец G
            for cell in row:
                if cell.value:
                    cell.font = Font(size=8)  # Устанавливаем размер шрифта 8

        # Выравнивание по центру для столбцов A, D, E, F, G, H, I
        center_alignment = Alignment(horizontal="center", vertical="center")
        center_columns = ['A', 'D', 'E', 'F', 'G', 'H', 'I']
        for col in center_columns:
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=columns.index(col) + 1, max_col=columns.index(col) + 1):
                for cell in row:
                    cell.alignment = center_alignment

    # Сохранение файла
    wb.save(output_path)

# def MK_creator(input_path, output_path):
#     xls = pd.ExcelFile(input_path)
#     wb = Workbook()
#     wb.remove(wb.active)  # Удаляем стандартный лист

#     sheet_counter = 1
#     row_limit = 18
#     current_row = 1
#     ws = None

#     for sheet_name in xls.sheet_names:
#         df = pd.read_excel(xls, sheet_name=sheet_name)

#         current_section = None
#         component_name = ""
#         quantity = None

#         for index, row in df.iterrows():
#             first_col = row.iloc[0] if len(row) > 0 else None  # Первая колонка (номер или пустая)
#             section = row.iloc[2] if len(row) > 2 else ""  # Название компонента или раздела
#             qty = row.iloc[3] if len(row) > 3 else None  # Количество

#             if ws is None:
#                 ws = wb.create_sheet(title=f"Лист{sheet_counter}")
#                 sheet_counter += 1
#                 current_row = 1

#             if pd.isna(first_col) and pd.isna(qty):  # Это раздел (название группы компонентов)
#                 if component_name:  # Добавляем предыдущий компонент в таблицу
#                     ws.append(["", component_name, f"{int(quantity)} шт."])
#                     current_row += 1
#                     component_name = ""

#                 current_section = " ".join(str(section).split())  # Убираем перенос строк

#                 if current_row > row_limit:
#                     ws = wb.create_sheet(title=f"Лист{sheet_counter}")
#                     sheet_counter += 1
#                     current_row = 1

#                 ws.append([current_section, "", ""])
#                 ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
#                 ws[current_row][0].font = Font(italic=True)
#                 ws[current_row][0].alignment = Alignment(horizontal="left")
#                 current_row += 1
#             else:  # Это компонент
#                 if not pd.isna(first_col):  # Начало нового компонента
#                     if component_name:  # Записываем предыдущий компонент
#                         ws.append(["", component_name, f"{int(quantity)} шт."])
#                         current_row += 1
#                     component_name = " ".join(str(section).split())  # Убираем перенос строк
#                     quantity = qty
#                 else:  # Продолжение названия компонента
#                     component_name += " " + " ".join(str(section).split())

#         if component_name:  # Добавляем последний компонент
#             ws.append(["", component_name, f"{int(quantity)} шт."])
#             current_row += 1

#     # Настройка ширины колонок
#     for sheet in wb.worksheets:
#         for col_num in range(1, 4):
#             col_letter = get_column_letter(col_num)
#             sheet.column_dimensions[col_letter].width = 30

#     wb.save(output_path)

def MK_cut_on_section(result, specification):
    section_names = specification[specification['Наименование'].str.contains('Конденсаторы|Микросхемы|Диоды|Транзисторы|Резисторы|Сборочные единицы', case=False, na=False)]['Наименование']
    final_data = []

    for _, row in result.iterrows():
        name = row['A']
        if name in section_names.values:
            final_data.append([name, '', ''])
        else:
            words = name.split()  # Разбиваем на слова
            lines = []
            current_line = ""

            for word in words:
                # if len(current_line) + len(word) + 1 <= 18:  # +1 для пробела
                current_line += (" " if current_line else "") + word
                # else:
                    # lines.append(current_line)
                    # current_line = word

            if current_line:
                lines.append(current_line)

            # Заполняем строки
            for i, line in enumerate(lines):
                if i == 0:
                    final_data.append([line, '',row['C']])
                else:
                    final_data.append([line, "", ""])  # Пустая строка с текстом
    return final_data

def MK_creator(input_path, res, output_path, specification):
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    row_count = 0
    sheet_number = 1

    italic_sections = [
        "Конденсаторы", "Микросхемы", "Катушки", "индуктивности", "Резисторы",
        "Печатная плата", "Транзисторы", "Диоды", "Соединения", "контактные", "Сборочные единицы", "Трансформаторы"
    ]
    result = pd.DataFrame({
        'A': res['Наименование'],
        'B': '',
        'C': res.get('Кол.', ''),
    })
    result['C'] = result['C'].apply(lambda x: str(int(x)) + " шт." if pd.notna(x) else "")
    final_data = MK_cut_on_section(result, specification)
    for row in final_data:
        if row_count >= 13:
            sheet_number += 1
            ws = wb.create_sheet(title=f"Лист{sheet_number}")
            row_count = 0

        ws.append(row)
        row_count += 1

    # Применение стилей ко всем листам
    for sheet in wb.worksheets:
        # Установка формата текста для столбцов G, H, I
        for col in sheet.iter_cols(min_col=7, max_col=9):
            for cell in col:
                cell.number_format = '@'

        # Установка ширины столбцов
        column_widths = [376, 376, 129]
        columns = ["A", "B", "C"]



        # Курсив для названий разделов
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=3):
            for cell in row:
                if cell.value:
                    cell.font = Font(size=12)  # Устанавливаем размер шрифта 8
                if cell.value and any(section in str(cell.value) for section in italic_sections):
                    cell.font = Font(size=12, italic=True, bold=True)

        # # Установка размера шрифта 8 для столбца G
        # for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=7, max_col=7):  # Столбец G
        #     for cell in row:
        #         if cell.value:
        #             cell.font = Font(size=8)  # Устанавливаем размер шрифта 8

        # Выравнивание по центру для столбцов A, D, E, F, G, H, I
        # center_alignment = Alignment(horizontal="center", vertical="center")
        # center_columns = ['A', 'D', 'E', 'F', 'G', 'H', 'I']
        # for col in center_columns:
        #     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=columns.index(col) + 1, max_col=columns.index(col) + 1):
        #         for cell in row:
        #             cell.alignment = center_alignment

    # Сохранение файла
    wb.save(output_path)





# def create_grouped_book(final_data, output_grouped_path):
    # """Создает книгу с компонентами, сгруппированными по категориям из результирующей таблицы."""
    # component_groups = [
    # "Конденсаторы", "Микросхемы", "Катушки индуктивности", "Резисторы", "Печатная плата",
    # "Транзисторы", "Диоды", "Соединения контактные"


def main():
    """Основная функция."""
    specification, passports = load_data(spec_file, ekb_file)
    specification, passports = prepare_data(specification, passports)
    specification = filter_unwanted_sections(specification)

    # Создание основной книги
    merged_data = merge_data(specification, passports)
    result = create_result_table(merged_data)
    final_data = add_section_names(result, specification)
    save_to_excel(final_data, output_path)

    # Создание книги с группировкой
    # create_grouped_book(final_data)


    # === Обработка ошибки PermissionError ===
    attempt = 0
    while attempt < 3:
        try:
            wb.save(output_grouped_path)
            print(f"Файл успешно сохранён: {output_grouped_path}")
            return
        except PermissionError:
            print(f"Ошибка: Файл {output_grouped_path} уже открыт. Закройте его и попробуйте снова.")
            time.sleep(3)  # Подождём перед повторной попыткой
            attempt += 1
        except FileNotFoundError:
            print(f"Ошибка: Папка для сохранения не существует! Проверьте путь: {os.path.dirname(output_grouped_path)}")
            return

    print("Не удалось сохранить файл. Проверьте права доступа или попробуйте изменить путь.")



def filter_unwanted_sections(specification):
    """Фильтрация ненужных разделов."""
    unwanted_sections = ["Документация", "Сборочный чертеж", "Сборочные единицы", "Плата печатная", "Прочие изделия", "Джамперы", "Оловянная перемычка"]

    # Убираем пробелы и приводим к нижнему регистру для точности сравнения
    specification['Наименование'] = specification['Наименование'].str.strip()

    # Применяем фильтрацию
    filtered_specification = specification[~specification['Наименование'].str.contains('|'.join(unwanted_sections), case=False, na=False)]

    return filtered_specification

def filter_unwanted_sections_MK(specification):
    """Фильтрация ненужных разделов."""
    unwanted_sections = ["Документация", "Сборочный чертеж", "Прочие изделия", "Джамперы", "Оловянная перемычка"]

    # Убираем пробелы и приводим к нижнему регистру для точности сравнения
    specification['Наименование'] = specification['Наименование'].str.strip()

    # Применяем фильтрацию
    filtered_specification = specification[~specification['Наименование'].str.contains('|'.join(unwanted_sections), case=False, na=False)]

    return filtered_specification

def main():
    """Основная функция."""
    specification, passports = load_data(spec_file, ekb_file)
    specification, passports = prepare_data(specification, passports)
    specification = filter_unwanted_sections(specification)
    merged_data = merge_data(specification, passports)
    result = create_result_table(merged_data)
    final_data = add_section_names(result, specification)
    save_to_excel(final_data, output_path)

if __name__ == "__main__":
    main()
